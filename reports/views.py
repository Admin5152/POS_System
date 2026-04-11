from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from accounts.decorators import role_required
from sales.models import Sale, SaleItem
from products.models import Product
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import timedelta, datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import io

@login_required
@role_required(['admin', 'manager', 'cashier'])
def dashboard(request):
    today = timezone.now().date()
    # Daily Sales
    daily_sales = Sale.objects.filter(date__date=today).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    # Weekly Sales
    week_ago = today - timedelta(days=7)
    weekly_sales = Sale.objects.filter(date__date__gte=week_ago).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Monthly Sales
    month_ago = today - timedelta(days=30)
    monthly_sales = Sale.objects.filter(date__date__gte=month_ago).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Recent Transactions
    recent_sales = Sale.objects.order_by('-date')[:10]
    
    return render(request, 'reports/dashboard.html', {
        'daily_sales': daily_sales,
        'weekly_sales': weekly_sales,
        'monthly_sales': monthly_sales,
        'recent_sales': recent_sales,
    })

@login_required
@role_required(['admin', 'manager'])
def sales_report(request):
    period = request.GET.get('period', 'daily')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
        end_date = today
        title = "Daily Sales Report"
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
        title = "Weekly Sales Report"
    elif period == 'monthly':
        start_date = today - timedelta(days=30)
        end_date = today
        title = "Monthly Sales Report"
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        title = f"Custom Sales Report ({start_date} to {end_date})"
    
    sales = Sale.objects.filter(date__date__gte=start_date, date__date__lte=end_date)
    
    # Calculate metrics
    total_transactions = sales.count()
    total_revenue = sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_tax = total_revenue * 0.1  # Assuming 10% tax rate
    total_discounts = 0  # You'll need to implement discount tracking
    net_sales = total_revenue - total_discounts
    
    # Employee performance
    employee_sales = sales.values('user__username').annotate(
        transactions=Count('id'),
        revenue=Sum('total_amount')
    ).order_by('-revenue')
    
    context = {
        'title': title,
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'total_transactions': total_transactions,
        'total_revenue': total_revenue,
        'total_tax': total_tax,
        'total_discounts': total_discounts,
        'net_sales': net_sales,
        'employee_sales': employee_sales,
        'sales': sales.order_by('-date')[:50]  # Last 50 transactions
    }
    
    return render(request, 'reports/sales_report.html', context)

@login_required
@role_required(['admin', 'manager'])
def product_sales_report(request):
    period = request.GET.get('period', 'monthly')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:  # monthly
        start_date = today - timedelta(days=30)
    
    # Get product sales data
    product_sales = SaleItem.objects.filter(
        sale__date__date__gte=start_date
    ).values(
        'product__product_name',
        'product__price'
    ).annotate(
        quantity_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('price'))
    ).order_by('-revenue')
    
    # Calculate totals for template
    total_product_revenue = sum(product['revenue'] for product in product_sales)
    avg_revenue_per_product = total_product_revenue / len(product_sales) if product_sales else 0
    
    context = {
        'title': f'Product Sales Report - {period.title()}',
        'period': period,
        'product_sales': product_sales,
        'total_product_revenue': total_product_revenue,
        'avg_revenue_per_product': avg_revenue_per_product,
        'start_date': start_date,
        'end_date': today
    }
    
    return render(request, 'reports/product_sales.html', context)

@login_required
@role_required(['admin', 'manager', 'cashier'])
def inventory_report(request):
    products = Product.objects.all().order_by('stock_quantity')
    low_stock_products = products.filter(stock_quantity__lte=10)
    
    # Calculate stock values
    products_with_value = []
    for product in products:
        product.stock_value = product.stock_quantity * product.price
        products_with_value.append(product)
    
    context = {
        'title': 'Inventory Report',
        'products': products_with_value,
        'low_stock_products': low_stock_products,
        'total_products': products.count(),
        'in_stock_count': products.count() - low_stock_products.count(),
        'low_stock_count': low_stock_products.count()
    }
    
    return render(request, 'reports/inventory_report.html', context)

@login_required
@role_required(['admin', 'manager'])
def cashier_performance_report(request):
    period = request.GET.get('period', 'monthly')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:  # monthly
        start_date = today - timedelta(days=30)
    
    # Get cashier performance data
    cashier_data = Sale.objects.filter(
        date__date__gte=start_date
    ).values('user__username').annotate(
        transactions=Count('id'),
        revenue=Sum('total_amount'),
        avg_transaction=Sum('total_amount') / Count('id')
    ).order_by('-revenue')
    
    # Calculate totals for template
    total_cashier_revenue = sum(cashier['revenue'] for cashier in cashier_data)
    avg_per_cashier = total_cashier_revenue / len(cashier_data) if cashier_data else 0
    
    context = {
        'title': f'Cashier Performance Report - {period.title()}',
        'period': period,
        'cashier_data': cashier_data,
        'total_cashier_revenue': total_cashier_revenue,
        'avg_per_cashier': avg_per_cashier,
        'start_date': start_date,
        'end_date': today
    }
    
    return render(request, 'reports/cashier_performance.html', context)

@login_required
@role_required(['admin', 'manager'])
def profit_report(request):
    period = request.GET.get('period', 'monthly')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
    else:  # monthly
        start_date = today - timedelta(days=30)
    
    # Calculate revenue
    sales = Sale.objects.filter(date__date__gte=start_date)
    total_revenue = sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate COGS
    sale_items = SaleItem.objects.filter(sale__date__date__gte=start_date)
    total_cogs = 0
    for item in sale_items:
        if item.product and item.product.cost_price:
            total_cogs += item.product.cost_price * item.quantity
    
    # Calculate profit
    gross_profit = total_revenue - total_cogs
    profit_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    context = {
        'title': f'Profit Report - {period.title()}',
        'period': period,
        'total_revenue': total_revenue,
        'total_cogs': total_cogs,
        'gross_profit': gross_profit,
        'profit_margin': profit_margin,
        'start_date': start_date,
        'end_date': today
    }
    
    return render(request, 'reports/profit_report.html', context)

# PDF Export Functions
def export_sales_pdf(request):
    period = request.GET.get('period', 'monthly')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
        end_date = today
        title = "Daily Sales Report"
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
        title = "Weekly Sales Report"
    else:
        start_date = today - timedelta(days=30)
        end_date = today
        title = "Monthly Sales Report"
    
    sales = Sale.objects.filter(date__date__gte=start_date, date__date__lte=end_date)
    total_transactions = sales.count()
    total_revenue = sales.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_tax = total_revenue * 0.1
    net_sales = total_revenue - total_tax
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{title}_{today}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center
    )
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 12))
    
    # Date range
    story.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Summary data
    summary_data = [
        ['Metric', 'Amount'],
        ['Total Transactions', str(total_transactions)],
        ['Total Revenue', f'${total_revenue:.2f}'],
        ['Total Tax (10%)', f'${total_tax:.2f}'],
        ['Net Sales', f'${net_sales:.2f}'],
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    doc.build(story)
    return response

def export_sales_excel(request):
    period = request.GET.get('period', 'monthly')
    today = timezone.now().date()
    
    if period == 'daily':
        start_date = today
        end_date = today
        title = "Daily Sales Report"
    elif period == 'weekly':
        start_date = today - timedelta(days=7)
        end_date = today
        title = "Weekly Sales Report"
    else:
        start_date = today - timedelta(days=30)
        end_date = today
        title = "Monthly Sales Report"
    
    sales = Sale.objects.filter(date__date__gte=start_date, date__date__lte=end_date)
    
    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    
    # Headers
    headers = ['Date', 'Transaction ID', 'Cashier', 'Payment Method', 'Total Amount']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Data
    for sale in sales:
        ws.append([
            sale.date.strftime('%Y-%m-%d %H:%M'),
            f"#{sale.id}",
            sale.user.username if sale.user else 'Unknown',
            sale.get_payment_method_display(),
            float(sale.total_amount)
        ])
    
    # Summary section
    ws.append([])
    ws.append(['SUMMARY'])
    ws.append(['Total Transactions', sales.count()])
    ws.append(['Total Revenue', f'${sales.aggregate(Sum("total_amount"))["total_amount__sum"] or 0:.2f}'])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{title}_{today}.xlsx"'
    wb.save(response)
    return response

@login_required
@role_required(['admin', 'manager', 'cashier'])
def modern_dashboard(request):
    """Modern dashboard view with beautiful UI"""
    try:
        from django.db.models import Sum, Count
        from sales.models import Sale
        from products.models import Product
        from customers.models import Customer
        from datetime import timedelta
        from django.utils import timezone
        
        today = timezone.now().date()
        
        # Calculate statistics
        total_sales = Sale.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_products = Product.objects.count()
        total_customers = Customer.objects.count()
        
        # Recent sales
        week_ago = today - timedelta(days=7)
        recent_sales = Sale.objects.filter(date__date__gte=week_ago).order_by('-date')[:10]
        
        # Sales by category
        sales_by_category = Sale.objects.all().values('items__product__category__name').annotate(
            category_sales=Sum('items__product__category__name')
        ).order_by('-category_sales')[:5]
        
        context = {
            'total_sales': total_sales,
            'total_products': total_products,
            'total_customers': total_customers,
            'recent_sales': recent_sales,
            'sales_by_category': list(sales_by_category),
        }
        
        return render(request, 'reports/modern_dashboard.html', context)
        
    except Exception as e:
        return render(request, 'reports/modern_dashboard.html', {
            'error': f'Unable to load dashboard: {str(e)}'
        })
